import traceback
import csv
import os
import json
import subprocess
import sys
import re
import shlex

from .process_papers import process_pubmed_id, save_output

from .workflow_context import WorkflowContext
from .model_data import ModelDataLoader
from .prompt_data import PromptDataParser

print("starting...")

model_data = ModelDataLoader()
parser = PromptDataParser()
context = WorkflowContext()


def setup(ctx=context):
    ctx.reinit(model_data)
    os.makedirs(ctx.data_cache_folder, exist_ok=True)
    os.makedirs(ctx.cache_folder, exist_ok=True)
    ctx.setup_llm_engine(model_data)


def read_pubmed_ids(file_path, column_name):
    """
    Reads PubMed IDs from a specified column in either a CSV or XLSX file.
    Returns a list of PubMed IDs.
    """
    pubmed_ids = []

    file_path = model_data.resolve_path(file_path)

    if file_path.endswith(".csv"):
        with open(file_path, mode="r", newline="", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            if column_name not in reader.fieldnames:
                raise ValueError(f"Column '{column_name}' not found in the CSV file.")

            pubmed_ids = [row[column_name] for row in reader if row[column_name]]

    elif file_path.endswith(".xlsx"):

        try:
            import openpyxl
        except ModuleNotFoundError as e:
            raise ValueError("To load Excel files openpyxl must be installed") from e

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # Find the column index by matching the header
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"Column '{column_name}' not found in the Excel file.")

        col_index = (
            headers.index(column_name) + 1
        )  # Convert to 1-based index for openpyxl

        # Extract PubMed IDs from the column
        for row in sheet.iter_rows(
            min_row=2, max_row=sheet.max_row, min_col=col_index, max_col=col_index
        ):
            if row[0].value:
                pubmed_ids.append(
                    str(row[0].value)
                )  # Convert to string to ensure consistency

    return pubmed_ids


def process_pubmed_ids(pubmed_ids, sections_to_extract, data_folder, ctx=context):
    processed_documents = []  # Store processed documents

    output_folder = model_data.get("output_folder", "output")
    output_file = model_data.get("output_file", "output.csv")

    output_file_base = output_file.split(".")[0]
    output_file_ext = output_file.split(".")[-1]

    output_file = os.path.join(output_folder, output_file)
    output_file_json = os.path.join(output_folder, output_file_base + ".json")

    debug_output_file = os.path.join(
        output_folder, output_file_base + "_debug." + output_file_ext
    )
    debug_output_file_json = os.path.join(
        output_folder, output_file_base + "_debug.json"
    )

    os.makedirs(output_folder, exist_ok=True)

    for pubmed_id in pubmed_ids[ctx.start_from :]:
        try:
            process_pubmed_id(
                pubmed_id,
                processed_documents,
                sections_to_extract,
                data_folder,
                ctx,
                model_data,
                parser,
            )
        except FileNotFoundError as e:
            print(f"Skipping {pubmed_id}: {e}.")
            continue
        except Exception as e:
            print(e)
            print("error with", pubmed_id)
            traceback.print_exc(file=sys.stdout)

        if ctx.max_docs is not None:
            if len(ctx.final_output) >= ctx.max_docs:
                break
        if len(ctx.final_output) > 0:
            save_output(ctx.final_output, output_file, output_file_json, ctx)
            save_output(ctx.debug, debug_output_file, debug_output_file_json, ctx)
        else:
            print("no output", pubmed_id)

    print("Final Output")
    print(ctx.final_output)
    print(ctx.ordered_column_list)

    # Save outputs
    if ctx.final_output:
        save_output(ctx.final_output, output_file, output_file_json, ctx)
    else:
        print(f"No final output to save to {output_file}")

    if ctx.debug:
        save_output(ctx.debug, debug_output_file, debug_output_file_json, ctx)
    else:
        print(f"No debug output to save to {debug_output_file}")

    return processed_documents


def search_for_pubmed_ids(search_script, search_term, search_options):
    pubmed_ids = []

    try:

        search_term = re.sub(r"([()])", r" \1 ", search_term).strip()

        search_term = search_term + " " + search_options
        # search_term = base64.b64encode(search_term.encode()).decode()

        args = shlex.split(search_term)
        result = subprocess.run(
            [search_script] + args, capture_output=True, text=True, check=True
        )
        # result = subprocess.run([search_script, "-b", search_term], capture_output=True, text=True, check=True)
        result_text = result.stdout

        try:
            # see if it is json
            pubmed_ids = json.loads(result_text)
            pubmed_ids = ["pmid"] + pubmed_ids
        except:
            pubmed_ids = result_text.splitlines()

    except subprocess.CalledProcessError as e:
        print("Error running search script")
        print(e)

    return pubmed_ids[1:], pubmed_ids[0]


def parse_papers(config_file, ctx=context):
    model_data.load_model_data(config_file)

    setup()

    parser.load_prompt_data(model_data)
    if ctx.use_pubmed_search:
        search_script = model_data.get("pubmed_search_script")
        search_term = model_data.get("pubmed_search_term")
        search_options = model_data.get("pubmed_search_options", "")

        pubmed_ids, ctx.column_name = search_for_pubmed_ids(
            search_script, search_term, search_options
        )
    else:
        file_path = model_data.get("documents_data")
        ctx.column_name = model_data.get("column_name")
        pubmed_ids = read_pubmed_ids(file_path, ctx.column_name)

    # Get the list of processed documents
    sections_to_extract = model_data.get("sections")
    num_pubmed_ids = len(pubmed_ids)

    print(f"Processing {num_pubmed_ids} documents.")

    processed_documents = process_pubmed_ids(
        pubmed_ids, sections_to_extract, ctx.data_cache_folder
    )
    print(f"Processed {len(processed_documents)} documents.")
